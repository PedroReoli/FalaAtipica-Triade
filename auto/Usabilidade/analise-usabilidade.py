"""
Script para análise de usabilidade - FalaAtípica
Processa dados do formulário HTML e gera planilha Excel com cálculos, gráficos e análise
"""

import os
import re
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg
import io

# Cores do projeto (sem # para openpyxl)
CORES = {
    'azul': '1e88e5',
    'verde': '43a047',
    'vermelho': 'e53935',
    'amarelo': 'fbc02d',
    'fundo_azul': '054776',
    'branco_geada': 'f4f6ff'
}

# Nomes das perguntas (baseado na estrutura do formulário)
PERGUNTAS = [
    "Usaria o sistema com frequência",
    "Sistema desnecessariamente complexo",
    "Fácil de usar",
    "Precisaria de suporte técnico",
    "Funcionalidades bem integradas",
    "Muita inconsistência na interface",
    "Aprenderia rapidamente",
    "Sistema confuso de usar",
    "Confiante ao usar",
    "Precisaria aprender muitas coisas"
]

def parse_html(html_file):
    """Extrai dados do arquivo HTML do Google Forms"""
    print(f"[INFO] Lendo arquivo: {html_file}")
    
    with open(html_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    soup = BeautifulSoup(content, 'html.parser')
    table = soup.find('table', class_='waffle')
    
    if not table:
        raise ValueError("Tabela não encontrada no HTML")
    
    rows = table.find_all('tr')
    data = []
    
    # Pular cabeçalho (linha 0) e linha de freeze
    for row in rows[2:]:
        cells = row.find_all('td')
        if len(cells) < 14:
            continue
        
        # Extrair dados de cada célula
        row_data = []
        for cell in cells:
            text = cell.get_text(strip=True)
            row_data.append(text)
        
        # Verificar se é uma linha de dados válida (tem data)
        if row_data[0] and re.match(r'\d{2}/\d{2}/\d{4}', row_data[0]):
            data.append(row_data)
    
    print(f"[OK] {len(data)} respostas encontradas")
    return data

def processar_dados(data):
    """Processa os dados brutos em DataFrame"""
    df = pd.DataFrame(data, columns=[
        'Data', 'P1', 'P2', 'P3', 'P4', 'P5', 'P6', 'P7', 'P8', 'P9', 'P10', 
        'Sugestao', 'Nome', 'Email'
    ])
    
    # Converter datas
    df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
    
    # Converter notas para numérico
    for i in range(1, 11):
        col = f'P{i}'
        df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Remover linhas com dados inválidos
    df = df.dropna(subset=['P1'])
    
    print(f"[OK] {len(df)} respostas válidas processadas")
    return df

def calcular_sus_score(df):
    """
    Calcula o SUS Score (System Usability Scale)
    SUS Score varia de 0 a 100, onde:
    - > 80: Excelente
    - 68-80: Bom
    - 51-67: OK
    - < 51: Precisa melhorias
    """
    # Perguntas ímpares (1, 3, 5, 7, 9) são positivas: (nota - 1) * 2.5
    # Perguntas pares (2, 4, 6, 8, 10) são negativas: (5 - nota) * 2.5
    sus_scores = []
    
    for idx, row in df.iterrows():
        score = 0
        
        # Perguntas positivas (ímpares)
        for i in [1, 3, 5, 7, 9]:
            col = f'P{i}'
            if pd.notna(row[col]):
                score += (row[col] - 1) * 2.5
        
        # Perguntas negativas (pares) - invertidas
        for i in [2, 4, 6, 8, 10]:
            col = f'P{i}'
            if pd.notna(row[col]):
                score += (5 - row[col]) * 2.5
        
        if score > 0:
            sus_scores.append(score)
    
    if sus_scores:
        return {
            'media': np.mean(sus_scores),
            'mediana': np.median(sus_scores),
            'desvio_padrao': np.std(sus_scores),
            'minimo': np.min(sus_scores),
            'maximo': np.max(sus_scores),
            'total': len(sus_scores),
            'classificacao': 'Excelente' if np.mean(sus_scores) > 80 else 
                           'Bom' if np.mean(sus_scores) >= 68 else 
                           'OK' if np.mean(sus_scores) >= 51 else 'Precisa Melhorias'
        }
    return None

def calcular_metricas_usabilidade(df, stats):
    """Calcula métricas específicas de usabilidade"""
    metricas = {}
    
    # 1. Satisfação Geral (média de todas as perguntas positivas)
    perguntas_positivas = ['P1', 'P3', 'P5', 'P7', 'P9']
    satisfacao_geral = []
    for col in perguntas_positivas:
        satisfacao_geral.extend(df[col].dropna().tolist())
    metricas['satisfacao_geral'] = {
        'media': np.mean(satisfacao_geral) if satisfacao_geral else 0,
        'total_respostas': len(satisfacao_geral)
    }
    
    # 2. Facilidade de Uso (P3)
    metricas['facilidade_uso'] = stats.get('Fácil de usar', {}).get('media', 0)
    
    # 3. Confiança (P9)
    metricas['confianca'] = stats.get('Confiante ao usar', {}).get('media', 0)
    
    # 4. Aprendizado (P7)
    metricas['aprendizado'] = stats.get('Aprenderia rapidamente', {}).get('media', 0)
    
    # 5. Complexidade Percebida (P2 - invertida)
    complexidade = stats.get('Sistema desnecessariamente complexo', {}).get('media', 0)
    metricas['simplicidade'] = 6 - complexidade  # Inverter escala
    
    # 6. Taxa de Satisfação (percentual de notas 4 e 5)
    notas_altas = sum(1 for v in satisfacao_geral if v >= 4)
    metricas['taxa_satisfacao'] = (notas_altas / len(satisfacao_geral) * 100) if satisfacao_geral else 0
    
    # 7. Taxa de Insatisfação (percentual de notas 1 e 2)
    notas_baixas = sum(1 for v in satisfacao_geral if v <= 2)
    metricas['taxa_insatisfacao'] = (notas_baixas / len(satisfacao_geral) * 100) if satisfacao_geral else 0
    
    return metricas

def calcular_estatisticas(df):
    """Calcula estatísticas descritivas"""
    stats = {}
    
    for i, pergunta in enumerate(PERGUNTAS, 1):
        col = f'P{i}'
        valores = df[col].dropna()
        
        if len(valores) > 0:
            stats[pergunta] = {
                'media': valores.mean(),
                'mediana': valores.median(),
                'desvio_padrao': valores.std(),
                'minimo': valores.min(),
                'maximo': valores.max(),
                'total_respostas': len(valores),
                'distribuicao': valores.value_counts().sort_index().to_dict()
            }
    
    return stats

def criar_grafico_barras(stats, pergunta, valores):
    """Cria gráfico de barras para uma pergunta"""
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Preparar dados
    labels = list(valores.keys())
    heights = list(valores.values())
    cores = ['#' + CORES['azul'] if h == max(heights) else '#' + CORES['verde'] for h in heights]
    
    # Criar gráfico
    bars = ax.bar(labels, heights, color=cores, edgecolor='white', linewidth=2)
    
    # Adicionar valores nas barras
    for bar in bars:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{int(height)}',
                ha='center', va='bottom', fontsize=12, fontweight='bold')
    
    ax.set_title(pergunta, fontsize=14, fontweight='bold', pad=20)
    ax.set_xlabel('Nota', fontsize=12)
    ax.set_ylabel('Quantidade de Respostas', fontsize=12)
    ax.set_ylim(0, max(heights) * 1.2)
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    
    plt.tight_layout()
    
    # Converter para bytes
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
    buf.seek(0)
    plt.close()
    
    return buf

def criar_grafico_medias(stats):
    """Cria gráfico de médias por pergunta"""
    perguntas = list(stats.keys())
    medias = [stats[p]['media'] for p in perguntas]
    
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Criar gráfico de barras horizontal
    y_pos = np.arange(len(perguntas))
    cores_barras = ['#' + CORES['azul'] if m >= 4 else '#' + CORES['verde'] if m >= 3 else '#' + CORES['amarelo'] for m in medias]
    
    bars = ax.barh(y_pos, medias, color=cores_barras, edgecolor='white', linewidth=2)
    
    # Adicionar valores
    for i, (bar, media) in enumerate(zip(bars, medias)):
        ax.text(media, bar.get_y() + bar.get_height()/2.,
                f' {media:.2f}',
                va='center', fontsize=10, fontweight='bold')
    
    ax.set_yticks(y_pos)
    ax.set_yticklabels([p[:50] + '...' if len(p) > 50 else p for p in perguntas], fontsize=9)
    ax.set_xlabel('Média das Notas', fontsize=12, fontweight='bold')
    ax.set_title('Média de Notas por Pergunta', fontsize=14, fontweight='bold', pad=20)
    ax.set_xlim(0, 5.5)
    ax.axvline(x=4, color='#' + CORES['verde'], linestyle='--', linewidth=2, label='Bom (≥4.0)')
    ax.axvline(x=3, color='#' + CORES['amarelo'], linestyle='--', linewidth=2, label='Regular (≥3.0)')
    ax.legend()
    ax.grid(axis='x', alpha=0.3, linestyle='--')
    
    plt.tight_layout()
    
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
    buf.seek(0)
    plt.close()
    
    return buf

def criar_planilha_excel(df, stats, sus_score, metricas, output_file):
    """Cria planilha Excel formatada com dados, cálculos e gráficos"""
    print(f"[INFO] Criando planilha Excel: {output_file}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Brutos"
    
    # Estilos
    header_fill = PatternFill(start_color=CORES['fundo_azul'], end_color=CORES['fundo_azul'], fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=11)  # FFFFFFFF = branco com alpha
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== ABA 1: DADOS BRUTOS ==========
    headers = ['Data', 'Nome', 'Email'] + PERGUNTAS + ['Sugestão']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados
    for row_num, (idx, row) in enumerate(df.iterrows(), 2):
        ws.cell(row=row_num, column=1, value=row['Data'].strftime('%d/%m/%Y %H:%M:%S') if pd.notna(row['Data']) else '')
        ws.cell(row=row_num, column=2, value=row['Nome'])
        ws.cell(row=row_num, column=3, value=row['Email'])
        
        for i, pergunta in enumerate(PERGUNTAS, 4):
            valor = row[f'P{i-3}']
            cell = ws.cell(row=row_num, column=i)
            cell.value = valor if pd.notna(valor) else ''
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        ws.cell(row=row_num, column=14, value=row['Sugestao'])
        ws.cell(row=row_num, column=14).border = border
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    for i in range(4, 14):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.column_dimensions['N'].width = 60
    
    # ========== ABA 2: ESTATÍSTICAS ==========
    ws2 = wb.create_sheet("Estatísticas")
    
    # Cabeçalho
    headers_stats = ['Pergunta', 'Média', 'Mediana', 'Desvio Padrão', 'Mínimo', 'Máximo', 'Total Respostas']
    for col_num, header in enumerate(headers_stats, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados estatísticos
    for row_num, (pergunta, stat) in enumerate(stats.items(), 2):
        ws2.cell(row=row_num, column=1, value=pergunta).border = border
        ws2.cell(row=row_num, column=2, value=round(stat['media'], 2)).border = border
        ws2.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_num, column=3, value=round(stat['mediana'], 2)).border = border
        ws2.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_num, column=4, value=round(stat['desvio_padrao'], 2)).border = border
        ws2.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_num, column=5, value=int(stat['minimo'])).border = border
        ws2.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_num, column=6, value=int(stat['maximo'])).border = border
        ws2.cell(row=row_num, column=6).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_num, column=7, value=int(stat['total_respostas'])).border = border
        ws2.cell(row=row_num, column=7).alignment = Alignment(horizontal='center')
    
    # Ajustar largura
    ws2.column_dimensions['A'].width = 50
    for i in range(2, 8):
        ws2.column_dimensions[get_column_letter(i)].width = 15
    
    # ========== ABA 3: DISTRIBUIÇÃO ==========
    ws3 = wb.create_sheet("Distribuição")
    
    # Cabeçalho
    headers_dist = ['Pergunta'] + [f'Nota {i}' for i in range(1, 6)]
    for col_num, header in enumerate(headers_dist, 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados de distribuição
    for row_num, (pergunta, stat) in enumerate(stats.items(), 2):
        ws3.cell(row=row_num, column=1, value=pergunta).border = border
        for nota in range(1, 6):
            valor = stat['distribuicao'].get(nota, 0)
            cell = ws3.cell(row=row_num, column=nota + 1, value=valor)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    ws3.column_dimensions['A'].width = 50
    for i in range(2, 7):
        ws3.column_dimensions[get_column_letter(i)].width = 12
    
    # ========== ABA 4: CÁLCULOS DE USABILIDADE ==========
    ws_calc = wb.create_sheet("Cálculos de Usabilidade")
    
    # Título
    title_cell = ws_calc['A1']
    title_cell.value = "CÁLCULOS DE USABILIDADE - FalaAtípica"
    title_cell.font = Font(bold=True, size=16, color='FF' + CORES['fundo_azul'])
    title_cell.alignment = Alignment(horizontal='center')
    ws_calc.merge_cells('A1:D1')
    
    # SUS Score
    row = 3
    ws_calc['A3'] = "SUS SCORE (System Usability Scale)"
    ws_calc['A3'].font = Font(bold=True, size=14, color='FF' + CORES['fundo_azul'])
    
    if sus_score:
        row += 1
        ws_calc['A4'] = "Métrica"
        ws_calc['B4'] = "Valor"
        ws_calc['A4'].font = Font(bold=True)
        ws_calc['B4'].font = Font(bold=True)
        
        row += 1
        ws_calc['A5'] = "SUS Score Médio"
        ws_calc['B5'] = round(sus_score['media'], 2)
        ws_calc['B5'].font = Font(bold=True, size=12)
        
        row += 1
        ws_calc['A6'] = "Classificação"
        ws_calc['B6'] = sus_score['classificacao']
        if sus_score['classificacao'] == 'Excelente':
            ws_calc['B6'].font = Font(bold=True, color='FF' + CORES['verde'])
        elif sus_score['classificacao'] == 'Bom':
            ws_calc['B6'].font = Font(bold=True, color='FF' + CORES['azul'])
        elif sus_score['classificacao'] == 'OK':
            ws_calc['B6'].font = Font(bold=True, color='FF' + CORES['amarelo'])
        else:
            ws_calc['B6'].font = Font(bold=True, color='FF' + CORES['vermelho'])
        
        row += 1
        ws_calc['A7'] = "Mediana"
        ws_calc['B7'] = round(sus_score['mediana'], 2)
        
        row += 1
        ws_calc['A8'] = "Desvio Padrão"
        ws_calc['B8'] = round(sus_score['desvio_padrao'], 2)
        
        row += 1
        ws_calc['A9'] = "Mínimo"
        ws_calc['B9'] = round(sus_score['minimo'], 2)
        
        row += 1
        ws_calc['A10'] = "Máximo"
        ws_calc['B10'] = round(sus_score['maximo'], 2)
        
        row += 1
        ws_calc['A11'] = "Total de Avaliações"
        ws_calc['B11'] = sus_score['total']
        
        # Interpretação SUS
        row += 3
        ws_calc[f'A{row}'] = "INTERPRETAÇÃO DO SUS SCORE:"
        ws_calc[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        ws_calc[f'A{row}'] = "• > 80: Excelente - Sistema de alta usabilidade"
        row += 1
        ws_calc[f'A{row}'] = "• 68-80: Bom - Sistema com boa usabilidade"
        row += 1
        ws_calc[f'A{row}'] = "• 51-67: OK - Sistema aceitável, mas pode melhorar"
        row += 1
        ws_calc[f'A{row}'] = "• < 51: Precisa Melhorias - Sistema com problemas de usabilidade"
    
    # Métricas de Usabilidade
    row += 3
    ws_calc[f'A{row}'] = "MÉTRICAS DE USABILIDADE"
    ws_calc[f'A{row}'].font = Font(bold=True, size=14, color='FF' + CORES['fundo_azul'])
    
    row += 1
    ws_calc[f'A{row}'] = "Métrica"
    ws_calc[f'B{row}'] = "Valor"
    ws_calc[f'C{row}'] = "Interpretação"
    for col in ['A', 'B', 'C']:
        ws_calc[f'{col}{row}'].font = Font(bold=True)
        ws_calc[f'{col}{row}'].fill = header_fill
        ws_calc[f'{col}{row}'].font = header_font
    
    row += 1
    ws_calc[f'A{row}'] = "Satisfação Geral"
    ws_calc[f'B{row}'] = f"{metricas['satisfacao_geral']['media']:.2f} / 5.0"
    if metricas['satisfacao_geral']['media'] >= 4.0:
        ws_calc[f'C{row}'] = "Excelente"
        ws_calc[f'C{row}'].font = Font(color='FF' + CORES['verde'], bold=True)
    elif metricas['satisfacao_geral']['media'] >= 3.0:
        ws_calc[f'C{row}'] = "Bom"
        ws_calc[f'C{row}'].font = Font(color='FF' + CORES['azul'], bold=True)
    else:
        ws_calc[f'C{row}'] = "Precisa Melhorias"
        ws_calc[f'C{row}'].font = Font(color='FF' + CORES['amarelo'], bold=True)
    
    row += 1
    ws_calc[f'A{row}'] = "Facilidade de Uso"
    ws_calc[f'B{row}'] = f"{metricas['facilidade_uso']:.2f} / 5.0"
    
    row += 1
    ws_calc[f'A{row}'] = "Confiança do Usuário"
    ws_calc[f'B{row}'] = f"{metricas['confianca']:.2f} / 5.0"
    
    row += 1
    ws_calc[f'A{row}'] = "Facilidade de Aprendizado"
    ws_calc[f'B{row}'] = f"{metricas['aprendizado']:.2f} / 5.0"
    
    row += 1
    ws_calc[f'A{row}'] = "Simplicidade Percebida"
    ws_calc[f'B{row}'] = f"{metricas['simplicidade']:.2f} / 5.0"
    
    row += 1
    ws_calc[f'A{row}'] = "Taxa de Satisfação"
    ws_calc[f'B{row}'] = f"{metricas['taxa_satisfacao']:.1f}%"
    ws_calc[f'B{row}'].font = Font(bold=True, color='FF' + CORES['verde'])
    
    row += 1
    ws_calc[f'A{row}'] = "Taxa de Insatisfação"
    ws_calc[f'B{row}'] = f"{metricas['taxa_insatisfacao']:.1f}%"
    if metricas['taxa_insatisfacao'] > 20:
        ws_calc[f'B{row}'].font = Font(bold=True, color='FF' + CORES['vermelho'])
    else:
        ws_calc[f'B{row}'].font = Font(bold=True, color='FF' + CORES['amarelo'])
    
    ws_calc.column_dimensions['A'].width = 35
    ws_calc.column_dimensions['B'].width = 20
    ws_calc.column_dimensions['C'].width = 30
    
    # ========== ABA 5: ANÁLISE ==========
    ws4 = wb.create_sheet("Análise")
    
    # Título
    title_cell = ws4['A1']
    title_cell.value = "ANÁLISE DE USABILIDADE - FalaAtípica"
    title_cell.font = Font(bold=True, size=16, color='FF' + CORES['fundo_azul'])  # Adicionar alpha
    title_cell.alignment = Alignment(horizontal='center')
    ws4.merge_cells('A1:F1')
    
    # Data de análise
    ws4['A3'] = f"Data de Análise: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws4['A3'].font = Font(size=11, italic=True)
    
    # Resumo geral
    ws4['A5'] = "RESUMO GERAL"
    ws4['A5'].font = Font(bold=True, size=12, color='FF' + CORES['fundo_azul'])
    
    total_respostas = len(df)
    media_geral = np.mean([stats[p]['media'] for p in stats.keys()])
    
    ws4['A6'] = f"Total de Respostas: {total_respostas}"
    ws4['A7'] = f"Média Geral: {media_geral:.2f}"
    ws4['A8'] = f"Desvio Padrão Geral: {np.std([stats[p]['media'] for p in stats.keys()]):.2f}"
    
    # Análise por pergunta
    ws4['A10'] = "ANÁLISE POR PERGUNTA"
    ws4['A10'].font = Font(bold=True, size=12, color='FF' + CORES['fundo_azul'])
    
    row = 11
    for pergunta, stat in stats.items():
        ws4[f'A{row}'] = pergunta
        ws4[f'A{row}'].font = Font(bold=True)
        
        media = stat['media']
        if media >= 4.0:
            avaliacao = "EXCELENTE"
            cor = 'FF' + CORES['verde']
        elif media >= 3.0:
            avaliacao = "BOM"
            cor = 'FF' + CORES['azul']
        elif media >= 2.0:
            avaliacao = "REGULAR"
            cor = 'FF' + CORES['amarelo']
        else:
            avaliacao = "PRECISA MELHORIAS"
            cor = 'FF' + CORES['vermelho']
        
        ws4[f'B{row}'] = f"Média: {media:.2f} - {avaliacao}"
        ws4[f'B{row}'].font = Font(color=cor, bold=True)
        
        # Observações
        if stat['desvio_padrao'] > 1.0:
            ws4[f'C{row}'] = "⚠ Alta variabilidade nas respostas"
            ws4[f'C{row}'].font = Font(italic=True, color='FF' + CORES['amarelo'])
        
        row += 1
    
    # Conclusões
    ws4[f'A{row + 2}'] = "CONCLUSÕES"
    ws4[f'A{row + 2}'].font = Font(bold=True, size=12, color='FF' + CORES['fundo_azul'])
    
    conclusoes = [
        f"O sistema recebeu {total_respostas} avaliações de usabilidade.",
        f"A média geral de {media_geral:.2f} indica uma avaliação {'positiva' if media_geral >= 4.0 else 'satisfatória' if media_geral >= 3.0 else 'que precisa melhorias'}.",
        "As perguntas com maior média indicam pontos fortes do sistema.",
        "As perguntas com menor média indicam áreas que precisam de atenção.",
        "Recomenda-se análise detalhada dos comentários para melhorias específicas."
    ]
    
    for i, conclusao in enumerate(conclusoes, row + 3):
        ws4[f'A{i}'] = f"• {conclusao}"
    
    ws4.column_dimensions['A'].width = 50
    ws4.column_dimensions['B'].width = 40
    ws4.column_dimensions['C'].width = 40
    
    # Melhorar análise com SUS Score
    if sus_score:
        row += 1
        ws4[f'A{row}'] = ""
        row += 1
        ws4[f'A{row}'] = "SUS SCORE"
        ws4[f'A{row}'].font = Font(bold=True, size=12, color='FF' + CORES['fundo_azul'])
        row += 1
        ws4[f'A{row}'] = f"Score Médio: {sus_score['media']:.2f} / 100"
        ws4[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        ws4[f'A{row}'] = f"Classificação: {sus_score['classificacao']}"
        if sus_score['classificacao'] == 'Excelente':
            ws4[f'A{row}'].font = Font(bold=True, color='FF' + CORES['verde'])
        elif sus_score['classificacao'] == 'Bom':
            ws4[f'A{row}'].font = Font(bold=True, color='FF' + CORES['azul'])
        row += 1
        ws4[f'A{row}'] = f"Interpretação: O SUS Score de {sus_score['media']:.2f} indica que o sistema possui usabilidade {'excelente' if sus_score['media'] > 80 else 'boa' if sus_score['media'] >= 68 else 'aceitável' if sus_score['media'] >= 51 else 'que precisa melhorias'}."
    
    # ========== ABA 6: GRÁFICOS ==========
    ws5 = wb.create_sheet("Gráficos")
    
    # Título
    title_cell = ws5['A1']
    title_cell.value = "GRÁFICOS DE USABILIDADE"
    title_cell.font = Font(bold=True, size=16, color='FF' + CORES['fundo_azul'])
    title_cell.alignment = Alignment(horizontal='center')
    ws5.merge_cells('A1:F1')
    
    # GRÁFICO 1: Médias por Pergunta
    row_start = 3
    ws5.cell(row=row_start, column=1, value="Pergunta").font = Font(bold=True)
    ws5.cell(row=row_start, column=2, value="Média").font = Font(bold=True)
    
    for row_num, (pergunta, stat) in enumerate(stats.items(), row_start + 1):
        ws5.cell(row=row_num, column=1, value=pergunta[:35])  # Truncar se muito longo
        ws5.cell(row=row_num, column=2, value=round(stat['media'], 2))
    
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Média de Notas por Pergunta"
    chart1.y_axis.title = "Média (1-5)"
    chart1.x_axis.title = "Pergunta"
    chart1.height = 12
    chart1.width = 18
    
    data = Reference(ws5, min_col=2, min_row=row_start + 1, max_row=row_start + len(stats))
    cats = Reference(ws5, min_col=1, min_row=row_start + 1, max_row=row_start + len(stats))
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.legend = None
    
    ws5.add_chart(chart1, "D3")
    
    # GRÁFICO 2: Distribuição de Notas (primeira pergunta)
    row_dist = None
    if stats:
        primeira_pergunta = list(stats.keys())[0]
        primeira_stat = stats[primeira_pergunta]
        
        row_dist = row_start + len(stats) + 5
        ws5.cell(row=row_dist, column=1, value="Nota").font = Font(bold=True)
        ws5.cell(row=row_dist, column=2, value="Quantidade").font = Font(bold=True)
        
        for i, nota in enumerate(range(1, 6), row_dist + 1):
            ws5.cell(row=i, column=1, value=nota)
            ws5.cell(row=i, column=2, value=primeira_stat['distribuicao'].get(nota, 0))
        
        chart2 = PieChart()
        chart2.title = f"Distribuição de Notas - {primeira_pergunta[:25]}"
        chart2.height = 10
        chart2.width = 10
        
        data2 = Reference(ws5, min_col=2, min_row=row_dist + 1, max_row=row_dist + 5)
        cats2 = Reference(ws5, min_col=1, min_row=row_dist + 1, max_row=row_dist + 5)
        chart2.add_data(data2, titles_from_data=False)
        chart2.set_categories(cats2)
        
        ws5.add_chart(chart2, "D18")
    
    # GRÁFICO 3: SUS Score (se disponível)
    row_sus = None
    if sus_score:
        row_sus = row_dist + 8 if row_dist else row_start + 15
        ws5.cell(row=row_sus, column=1, value="Métrica").font = Font(bold=True)
        ws5.cell(row=row_sus, column=2, value="Valor").font = Font(bold=True)
        
        ws5.cell(row=row_sus + 1, column=1, value="SUS Score")
        ws5.cell(row=row_sus + 1, column=2, value=round(sus_score['media'], 2))
        
        chart3 = BarChart()
        chart3.type = "col"
        chart3.style = 10
        chart3.title = "SUS Score - System Usability Scale"
        chart3.y_axis.title = "Score (0-100)"
        chart3.height = 8
        chart3.width = 12
        
        data3 = Reference(ws5, min_col=2, min_row=row_sus + 1, max_row=row_sus + 1)
        chart3.add_data(data3, titles_from_data=False)
        chart3.y_axis.scaling.min = 0
        chart3.y_axis.scaling.max = 100
        
        ws5.add_chart(chart3, "D30")
    
    # GRÁFICO 4: Taxa de Satisfação vs Insatisfação
    if row_sus:
        row_taxa = row_sus + 4
    elif row_dist:
        row_taxa = row_dist + 8
    else:
        row_taxa = row_start + 15
    ws5.cell(row=row_taxa, column=1, value="Categoria").font = Font(bold=True)
    ws5.cell(row=row_taxa, column=2, value="Percentual").font = Font(bold=True)
    
    ws5.cell(row=row_taxa + 1, column=1, value="Satisfeitos (4-5)")
    ws5.cell(row=row_taxa + 1, column=2, value=round(metricas['taxa_satisfacao'], 1))
    
    ws5.cell(row=row_taxa + 2, column=1, value="Neutros (3)")
    neutros = 100 - metricas['taxa_satisfacao'] - metricas['taxa_insatisfacao']
    ws5.cell(row=row_taxa + 2, column=2, value=round(neutros, 1))
    
    ws5.cell(row=row_taxa + 3, column=1, value="Insatisfeitos (1-2)")
    ws5.cell(row=row_taxa + 3, column=2, value=round(metricas['taxa_insatisfacao'], 1))
    
    chart4 = PieChart()
    chart4.title = "Distribuição de Satisfação"
    chart4.height = 10
    chart4.width = 10
    
    data4 = Reference(ws5, min_col=2, min_row=row_taxa + 1, max_row=row_taxa + 3)
    cats4 = Reference(ws5, min_col=1, min_row=row_taxa + 1, max_row=row_taxa + 3)
    chart4.add_data(data4, titles_from_data=False)
    chart4.set_categories(cats4)
    
    ws5.add_chart(chart4, "K18")
    
    # Salvar
    wb.save(output_file)
    print(f"[OK] Planilha Excel criada com sucesso: {output_file}")

def main():
    """Função principal"""
    print("=" * 60)
    print("  ANÁLISE DE USABILIDADE - FalaAtípica")
    print("=" * 60)
    print()
    
    # Caminhos
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(script_dir)
    
    # Procurar arquivo HTML
    html_file = None
    possiveis_nomes = [
        "Respostas ao formulário 1.html",
        "Respostas ao formulario 1.html",
        "respostas-ao-formulario-1.html"
    ]
    
    for nome in possiveis_nomes:
        caminho = os.path.join(root_dir, nome)
        if os.path.exists(caminho):
            html_file = caminho
            break
    
    if not html_file:
        # Tentar encontrar qualquer arquivo HTML na raiz
        for arquivo in os.listdir(root_dir):
            if arquivo.endswith('.html') and 'formulario' in arquivo.lower():
                html_file = os.path.join(root_dir, arquivo)
                print(f"[INFO] Arquivo HTML encontrado: {arquivo}")
                break
    
    if not html_file:
        raise FileNotFoundError(
            f"Arquivo HTML não encontrado na raiz do projeto.\n"
            f"Procurando em: {root_dir}\n"
            f"Certifique-se de que o arquivo HTML do Google Forms está na raiz do projeto."
        )
    
    output_file = os.path.join(root_dir, "Analise_Usabilidade.xlsx")
    
    try:
        # 1. Parsear HTML
        data = parse_html(html_file)
        
        # 2. Processar dados
        df = processar_dados(data)
        
        # 3. Calcular estatísticas
        print("[INFO] Calculando estatísticas...")
        stats = calcular_estatisticas(df)
        print(f"[OK] Estatísticas calculadas para {len(stats)} perguntas")
        
        # 4. Calcular SUS Score
        print("[INFO] Calculando SUS Score...")
        sus_score = calcular_sus_score(df)
        if sus_score:
            print(f"[OK] SUS Score: {sus_score['media']:.2f} ({sus_score['classificacao']})")
        
        # 5. Calcular métricas de usabilidade
        print("[INFO] Calculando métricas de usabilidade...")
        metricas = calcular_metricas_usabilidade(df, stats)
        print(f"[OK] Métricas calculadas")
        
        # 6. Criar planilha Excel
        criar_planilha_excel(df, stats, sus_score, metricas, output_file)
        
        # 5. Resumo final
        print()
        print("=" * 60)
        print("  RESUMO DA ANÁLISE")
        print("=" * 60)
        print(f"Total de Respostas: {len(df)}")
        print(f"Total de Perguntas: {len(stats)}")
        media_geral = np.mean([stats[p]['media'] for p in stats.keys()])
        print(f"Média Geral: {media_geral:.2f}")
        print()
        print(f"Planilha gerada: {output_file}")
        print("=" * 60)
        
    except Exception as e:
        print(f"[ERRO] Erro durante processamento: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())

